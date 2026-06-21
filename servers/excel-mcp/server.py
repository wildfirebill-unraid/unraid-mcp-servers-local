import json
import os
import csv
import io
from pathlib import Path

from mcp.server import Server, stdio_server
from mcp.types import Tool, TextContent, Resource
import anyio
import openpyxl


def _resolve_path(base: str, requested: str) -> str:
    p = (Path(base) / requested).resolve()
    if base:
        base_resolved = Path(base).resolve()
        if base_resolved not in p.parents and p != base_resolved:
            raise PermissionError(f"Access denied: {requested} is outside base path")
    if not p.exists():
        return str(p)
    return str(p)


class ExcelServer(Server):
    def __init__(self):
        super().__init__("excel")
        self._init_env()

    def _init_env(self):
        self._excel_path = os.environ.get("EXCEL_PATH", "")

    async def list_tools(self) -> list[Tool]:
        return [
            Tool(name="read_sheet", description="Read all data from a sheet as list of rows",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "Sheet name"}
                 }, "required": ["path", "sheet"]}),
            Tool(name="list_sheets", description="List all sheet names in workbook",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"}
                 }, "required": ["path"]}),
            Tool(name="write_cell", description="Write a value to a specific cell",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "Sheet name"},
                     "cell": {"type": "string", "description": "Cell reference (e.g. A1)"},
                     "value": {"type": "string", "description": "Value to write"}
                 }, "required": ["path", "sheet", "cell", "value"]}),
            Tool(name="read_cell", description="Read value from a specific cell",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "Sheet name"},
                     "cell": {"type": "string", "description": "Cell reference (e.g. A1)"}
                 }, "required": ["path", "sheet", "cell"]}),
            Tool(name="create_sheet", description="Create a new sheet in workbook",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "New sheet name"}
                 }, "required": ["path", "sheet"]}),
            Tool(name="sheet_info", description="Get dimensions, row count, column count of a sheet",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "Sheet name"}
                 }, "required": ["path", "sheet"]}),
            Tool(name="read_range", description="Read a range of cells from a sheet",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "Sheet name"},
                     "min_row": {"type": "integer", "description": "Minimum row (1-indexed)"},
                     "max_row": {"type": "integer", "description": "Maximum row (1-indexed)"},
                     "min_col": {"type": "integer", "description": "Minimum column (1-indexed)"},
                     "max_col": {"type": "integer", "description": "Maximum column (1-indexed)"}
                 }, "required": ["path", "sheet", "min_row", "max_row", "min_col", "max_col"]}),
            Tool(name="excel_to_csv", description="Convert a sheet to CSV string",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path to excel file"},
                     "sheet": {"type": "string", "description": "Sheet name"}
                 }, "required": ["path", "sheet"]}),
            Tool(name="create_workbook", description="Create a new empty workbook",
                 inputSchema={"type": "object", "properties": {
                     "path": {"type": "string", "description": "Path for new workbook"}
                 }, "required": ["path"]}),
        ]

    async def call_tool(self, name: str, arguments: dict | None) -> list[TextContent]:
        args = arguments or {}
        try:
            if name == "read_sheet":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[args["sheet"]]
                rows = [[c.value for c in row] for row in ws.iter_rows()]
                wb.close()
                return [TextContent(type="text", text=json.dumps(rows))]
            if name == "list_sheets":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                return [TextContent(type="text", text=json.dumps(sheets))]
            if name == "write_cell":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path)
                ws = wb[args["sheet"]]
                ws[args["cell"]] = args["value"]
                wb.save(path)
                wb.close()
                return [TextContent(type="text", text=json.dumps({"status": "ok"}))]
            if name == "read_cell":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[args["sheet"]]
                value = ws[args["cell"]].value
                wb.close()
                return [TextContent(type="text", text=json.dumps({"value": value}))]
            if name == "create_sheet":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path)
                wb.create_sheet(args["sheet"])
                wb.save(path)
                wb.close()
                return [TextContent(type="text", text=json.dumps({"status": "ok"}))]
            if name == "sheet_info":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[args["sheet"]]
                dim = ws.dimensions
                info = {"dimensions": dim, "rows": ws.max_row, "cols": ws.max_column}
                wb.close()
                return [TextContent(type="text", text=json.dumps(info))]
            if name == "read_range":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[args["sheet"]]
                rows = []
                for r in ws.iter_rows(min_row=args["min_row"], max_row=args["max_row"],
                                      min_col=args["min_col"], max_col=args["max_col"]):
                    rows.append([c.value for c in r])
                wb.close()
                return [TextContent(type="text", text=json.dumps(rows))]
            if name == "excel_to_csv":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[args["sheet"]]
                buf = io.StringIO()
                w = csv.writer(buf)
                for row in ws.iter_rows():
                    w.writerow([c.value for c in row])
                wb.close()
                return [TextContent(type="text", text=buf.getvalue())]
            if name == "create_workbook":
                path = _resolve_path(self._excel_path, args["path"])
                wb = openpyxl.Workbook()
                wb.save(path)
                wb.close()
                return [TextContent(type="text", text=json.dumps({"status": "ok", "path": path}))]
            raise ValueError(f"Unknown tool: {name}")
        except Exception as e:
            return [TextContent(type="text", text=json.dumps({"error": str(e)}))]

    async def list_resources(self) -> list[Resource]:
        return []

    async def read_resource(self, uri: str) -> str:
        raise ValueError(f"Unknown resource: {uri}")


async def main():
    server = ExcelServer()
    async with stdio_server() as (read, write):
        await server.run(read, write, server.list_tools, server.call_tool, server.list_resources, server.read_resource)

if __name__ == "__main__":
    anyio.run(main)
